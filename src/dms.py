import json
import os
import re
import sys
import textwrap
from datetime import datetime

import boto3
import openpyxl
from tabulate import tabulate

from config import (DB_LOG_FILE_COUNT, MAX_TASKS_PER_PAGE, SOURCE_DB_ID,
                    TARGET_DB_ID, json_files_location,
                    replication_instance_arn, sns_topic_arn,
                    source_endpoint_arn, target_endpoint_arn, task_arn_file)
from databases.oracle import oracle_table_metadata
from databases.postgres import postgres_table_metadata
from task_settings import task_settings


def create_dms_tasks(profile, region):
    """
    Reads all the json files and generates DMS tasks
    """

    session = boto3.Session(profile_name=profile, region_name=region)
    dms = session.client("dms")

    arn_list = []
    count = 0

    # Create tasks using JSON files
    for json_file in os.listdir(json_files_location):
        count += 1

        file_handler = open(os.path.join(json_files_location, json_file), "r")
        table_mapping = json.dumps((json.load(file_handler)))
        file_handler.close()

        task_id = json_file

        current_time = (
            datetime.now()
            .strftime("%Y-%m-%d %H:%M")
            .replace(" ", "-")
            .replace(":", "-")
        )

        # Replace special chars, otherwise AWS will complain.
        task_id = (
            task_id.replace(".json", "").replace("_", "-").replace(".", "-").strip()
            + "-"
            + current_time
        )

        try:
            response = dms.create_replication_task(
                ReplicationTaskIdentifier=task_id,
                SourceEndpointArn=source_endpoint_arn,
                TargetEndpointArn=target_endpoint_arn,
                ReplicationInstanceArn=replication_instance_arn,
                MigrationType="full-load",
                TableMappings=table_mapping,
                ReplicationTaskSettings=task_settings,
            )

            task_arn = response["ReplicationTask"]["ReplicationTaskArn"]

            print(f"{count} - DMS task created for file: {json_file}")
            arn_list.append(task_arn)

        except Exception as err:
            print(f"Error creating DMS task for file: {json_file}")
            print(err)
            print(
                "\nNOTE: Are you sure you have the correct AWS profile? Check the '--profile' paramter."
            )
            print(
                "      If no profile is passed, [default] profile will be used. It may not have permission to create a DMS task!!"
            )
            sys.exit(1)

    # Wait for the tasks to be in "READY" state
    wait_for_status_change(dms, "replication_task_ready", arn_list)

    # Persist the ARNs in a file.
    with open(task_arn_file, "w") as file_handle:
        [file_handle.write("%s\n" % arn) for arn in arn_list]

    print(f"{len(arn_list)} tasks have been created and ready")


def wait_for_status_change(dms, waiter_state, arn_list):
    """
    Creates waiters for DMS.
    """
    waiter = dms.get_waiter(waiter_state)

    waiter.wait(
        Filters=[
            {"Name": "replication-task-arn", "Values": arn_list},
        ],
    )


def list_dms_tasks(profile, region):
    """
    Prints the list of DMS tasks
    """
    session = boto3.Session(profile_name=profile, region_name=region)
    dms = session.client("dms")

    try:
        response = dms.describe_replication_tasks(
            MaxRecords=MAX_TASKS_PER_PAGE,
            Filters=[
                {
                    "Name": "endpoint-arn",
                    "Values": [
                        source_endpoint_arn,
                        target_endpoint_arn,
                    ],
                }
            ],
        )

    except Exception as err:
        print("Something went wrong while listing DMS tasks")
        print(err)
        sys.exit(1)

    tasks = []

    for task in response["ReplicationTasks"]:
        err_msg = ""
        start_date = ""

        if "LastFailureMessage" in task.keys():
            err_msg = task["LastFailureMessage"]

        if "StartDate" in task["ReplicationTaskStats"].keys():
            start_date = task["ReplicationTaskStats"]["StartDate"].strftime(
                "%Y-%m-%d %H:%M"
            )

        tasks.append(
            [
                task["ReplicationTaskIdentifier"],
                task["ReplicationTaskArn"],
                task["Status"],
                start_date,
                err_msg,
            ]
        )

    print(
        tabulate(
            tasks,
            headers=["Task ID", "Task ARN", "Status", "Start Date", "Error Message"],
            tablefmt="fancy_grid",
        )
    )


def start_dms_tasks(profile, region):
    """
    Starts the DMS tasks.

    Tasks must have been created before calling this function. It reads the
    "task_arn_file" and starts the tasks.
    """
    session = boto3.Session(profile_name=profile, region_name=region)
    dms = session.client("dms")

    count = 0
    task_arn_list = []

    # Read the task ARNs from "task_arn.txt" file and start the DMS tasks.
    with open(task_arn_file, "r") as arn_file:
        for task_arn in arn_file:
            task_arn = task_arn.strip("\n")
            task_arn_list.append(task_arn)

            try:
                response = dms.start_replication_task(
                    ReplicationTaskArn=task_arn,
                    StartReplicationTaskType="reload-target",
                )
                print("Task: {} has been started".format(task_arn))
            except Exception as error:
                count += 1
                print("Error starting task with ARN: {}".format(task_arn))
                print(error)

    # Once all the tasks have been started, we wanted to wait for all of them
    # to get completed. that's when their status change to 'replication_task_stopped'.
    # However, there seems to be a bug with this waiter in boto3.
    # https://github.com/boto/boto3/issues/1926
    # wait_for_status_change('replication_task_stopped', arn_list)

    # So, we are simply starting the tasks and return. we are NOT waiting for them
    # to get completed.
    if count > 0:
        msg = f"{count} errors encountered while starting DMS tasks."
    else:
        msg = f"{len(task_arn_list)} tasks have been started"

    send_mail(profile, region, msg)


def delete_dms_tasks(profile, region):
    """
    Delete DMS tasks. The tasks to be deleted come from "task_arn.txt" file.
    """
    session = boto3.Session(profile_name=profile, region_name=region)
    dms = session.client("dms")

    count = 0
    arns_to_be_deleted = []

    with open(task_arn_file, "r") as arn_file:
        for arn in arn_file:
            arn = arn.strip("\n")
            arns_to_be_deleted.append(arn)

            try:
                response = dms.delete_replication_task(ReplicationTaskArn=arn)
                print("Task: {} deletion in progress...".format(arn))
            except Exception as error:
                count += 1
                print("Error deleting task with ARN: {}".format(arn))
                print(error)

    if count > 0:
        print(f"{count} errors encountered while deleting DMS tasks.")
    else:
        wait_for_status_change(dms, "replication_task_deleted", arns_to_be_deleted)
        print(f"{len(arns_to_be_deleted)} tasks have been deleted!")


def send_mail(profile, region, message):
    session = boto3.Session(profile_name=profile, region_name=region)
    sns = session.client("sns")

    try:
        if len(sns_topic_arn) > 0:
            sns.publish(
                TopicArn=sns_topic_arn,
                Message=message,
            )
    except Exception as exception:
        print("Error sending email: {}".format(exception))


def test_db_connection(profile, region):
    """
    Tests the connection between the replication instance and the endpoint.
    """
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        dms = session.client("dms")

        result = []

        def test_connection(instance_arn, db_endpoint_arn):
            # Test Source DB connection from the replication instance
            response = dms.test_connection(
                ReplicationInstanceArn=instance_arn,
                EndpointArn=db_endpoint_arn,
            )

            waiter = dms.get_waiter("test_connection_succeeds")
            waiter.wait()

            return [
                response["Connection"]["ReplicationInstanceArn"],
                response["Connection"]["ReplicationInstanceIdentifier"],
                response["Connection"]["EndpointArn"],
                response["Connection"]["EndpointIdentifier"],
                "Successful",
            ]

        # Test Source DB connection from the replication instance
        # Source DB Connection
        status = test_connection(replication_instance_arn, source_endpoint_arn)
        result.append(status)

        # Target DB Connection
        status = test_connection(replication_instance_arn, target_endpoint_arn)
        result.append(status)

        print(
            tabulate(
                result,
                headers=[
                    "Replication Instance ARN",
                    "Instance ID",
                    "Endpoint ARN",
                    "Endpoint ID",
                    "Status",
                ],
                tablefmt="fancy_grid",
            )
        )

    except Exception as err:
        print("Something went wrong while testing the connection")
        print(err)
        sys.exit(1)


def describe_table_statistics(profile, region):
    """
    Describe Table Statistics
    """
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        dms = session.client("dms")

        result = []

        with open(task_arn_file, "r") as arn_file:
            for task_arn in arn_file:
                task_arn = task_arn.strip("\n")

            response = dms.describe_table_statistics(
                ReplicationTaskArn=task_arn,
                MaxRecords=123,
                # Marker='string',
                # Filters=[
                #     {
                #         'Name': 'string',
                #         'Values': [
                #             'string',
                #         ]
                #     },
                # ]
            )

            for table_statistics in response["TableStatistics"]:
                result.append(
                    [
                        task_arn,
                        table_statistics["SchemaName"],
                        table_statistics["TableName"],
                        table_statistics["TableState"],
                        table_statistics["Inserts"],
                        table_statistics["Updates"],
                        table_statistics["Deletes"],
                        table_statistics["FullLoadRows"],
                        table_statistics["FullLoadErrorRows"],
                        table_statistics["FullLoadStartTime"].strftime(
                            "%Y-%m-%d %H:%M"
                        ),
                        table_statistics["FullLoadEndTime"].strftime("%Y-%m-%d %H:%M"),
                    ]
                )

        result.sort(key=lambda x: x[1] + x[2])

        header = [
            "Task ARN",
            "Schema",
            "Table",
            "State",
            "Inserts",
            "Updates",
            "Deletes",
            "Full Load Rows",
            "Full Load Error Rows",
            "Full Load Start Time",
            "Full Load End Time",
        ]
        print(tabulate(result, headers=header, tablefmt="fancy_grid"))

    except Exception as error:
        print("** Something went wrong while describing table statistics. **")
        print(error)

    # Fetch Source & Target DB Logs
    print("\n")
    print("*" * 120)
    print("Source & Target DB latest DB logs")
    print("*" * 120)
    describe_db_log_files(profile, region)


def create_iam_role_for_dms_cloudwatch_logs(profile, region):
    """
    Create IAM role for DMS CloudWatch Logs

    It seems, to create CloudWatch logs, AWS DMS needs to have a role with this
    name: "dms-cloudwatch-logs-role". This role needs to have the following
    policies attached: "AmazonDMSCloudWatchLogsRole

    Link:
    https://aws.amazon.com/premiumsupport/knowledge-center/dms-cloudwatch-logs-not-appearing
    """
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        iam = session.client("iam")

        response = iam.create_role(
            RoleName="dms-cloudwatch-logs-role",
            AssumeRolePolicyDocument=json.dumps(
                {
                    "Version": "2012-10-17",
                    "Statement": [
                        {
                            "Effect": "Allow",
                            "Principal": {"Service": "dms.amazonaws.com"},
                            "Action": "sts:AssumeRole",
                        }
                    ],
                }
            ),
        )

        response = iam.attach_role_policy(
            RoleName="dms-cloudwatch-logs-role",
            PolicyArn="arn:aws:iam::aws:policy/service-role/AmazonDMSCloudWatchLogsRole",
        )

        print("Created IAM role for DMS CloudWatch Logs")
    except Exception as error:
        print(
            "** Something went wrong while creating IAM role for DMS CloudWatch Logs. **"
        )
        print(error)


def fetch_cloudwatch_logs_for_a_task(profile, region, task_arn):
    """
    Fetch CloudWatch Logs for a Task

    Input:
        Task ARN
    """
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        dms = session.client("dms")
        cloudwatch = session.client("logs")

        response = dms.describe_replication_tasks(
            Filters=[
                {
                    "Name": "replication-task-arn",
                    "Values": [
                        task_arn,
                    ],
                }
            ]
        )

        replication_task = response["ReplicationTasks"][0]

        cloudwatch_log_group = json.loads(replication_task["ReplicationTaskSettings"])[
            "Logging"
        ]["CloudWatchLogGroup"]
        cloudwatch_log_stream = json.loads(replication_task["ReplicationTaskSettings"])[
            "Logging"
        ]["CloudWatchLogStream"]

        response = cloudwatch.filter_log_events(
            logGroupName=cloudwatch_log_group,
            logStreamNames=[
                cloudwatch_log_stream,
            ],
            # logStreamNamePrefix='string',
            # startTime=123,
            # endTime=123,
            # filterPattern='string',
            # nextToken='string',
            # limit=123,
            # interleaved=True|False
        )

        print(f"Log Group : {cloudwatch_log_group}")
        print(f"Log Stream: {cloudwatch_log_stream}")

        result = []

        for line, event in enumerate(response["events"]):
            result.append(
                [
                    line,
                    event["message"][0:16],
                    event["message"].split("]")[1][0],
                    re.search(r"\[.*\]", event["message"][0:50]).group(0),
                    "\n".join(
                        textwrap.wrap(
                            event["message"][49:], width=150, replace_whitespace=False
                        )
                    ),
                ]
            )

        print(
            tabulate(
                result,
                headers=[
                    "line",
                    "Timestamp",
                    "Message Type",
                    "Source",
                    "Message",
                ],
                tablefmt="fancy_grid",
            )
        )

        # print(json.dumps(response, indent = 4, sort_keys=True, default=str))

    except Exception as error:
        print("** Something went wrong while fetching CloudWatch Logs for a Task. **")
        print(error)


def describe_endpoints(profile, region, print_result=True):
    """ """
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        dms = session.client("dms")

        result = []

        response = dms.describe_endpoints(
            Filters=[
                {
                    "Name": "endpoint-arn",
                    "Values": [
                        source_endpoint_arn,
                        target_endpoint_arn,
                    ],
                },
            ],
        )

        for db_endpoint in response["Endpoints"]:
            db_specific_key = db_endpoint["EngineDisplayName"] + "Settings"

            extra_connection_attributes = ""
            if "ExtraConnectionAttributes" in db_endpoint.keys():
                extra_connection_attributes = db_endpoint["ExtraConnectionAttributes"]

            result.append(
                [
                    db_endpoint["EndpointIdentifier"],
                    db_endpoint["EndpointType"],
                    db_endpoint["EngineDisplayName"],
                    db_endpoint[db_specific_key]["ServerName"],
                    db_endpoint[db_specific_key]["DatabaseName"],
                    db_endpoint[db_specific_key]["Port"],
                    db_endpoint[db_specific_key]["Username"],
                    extra_connection_attributes,
                ]
            )

        if print_result:
            print(
                tabulate(
                    result,
                    headers=[
                        "Endpoint_ID",
                        "Type",
                        "Database",
                        "Server",
                        "DB",
                        "Port",
                        "User",
                        "Extra Attributes",
                    ],
                    tablefmt="fancy_grid",
                )
            )

        # print(json.dumps(response, indent=4, sort_keys=True, default=str))
        return result

    except Exception as err:
        print("** Something went wrong while describing DB Endpoints **")
        print(err)
        sys.exit(1)


def describe_db_log_files(profile, region):
    """
    Get DB Logs

    Input:
        Profile
        Region
    """
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        rds = session.client("rds")

        def fetch_log_file(db_id):
            result = []

            try:
                response = rds.describe_db_log_files(
                    DBInstanceIdentifier=db_id,
                )

                for log_file in response["DescribeDBLogFiles"][
                    : -1 - 1 * DB_LOG_FILE_COUNT : -1
                ]:
                    resp = rds.download_db_log_file_portion(
                        DBInstanceIdentifier=db_id,
                        LogFileName=log_file["LogFileName"],
                    )

                    for line in resp["LogFileData"].split("\n"):
                        result.append(
                            [
                                db_id,
                                log_file["LogFileName"],
                                "\n".join(
                                    textwrap.wrap(
                                        line, width=150, replace_whitespace=False
                                    ),
                                ),
                            ],
                        )
            except Exception as error:
                print(
                    f"** Something went wrong while fetching DB Logs for DB Instance: {db_id} **"
                )
                print("IS THE DB INSTANCE NAME CORRECT ??")
                print(error)

            return result

        logs = fetch_log_file(SOURCE_DB_ID)
        print(tabulate(logs, headers=["DB_ID", "Logs"], tablefmt="fancy_grid"))

        logs = fetch_log_file(TARGET_DB_ID)
        print(tabulate(logs, headers=["DB_ID", "Logs"], tablefmt="fancy_grid"))

    except Exception as error:
        print("** Something went wrong while getting DB Logs. **")
        print(error)


def get_source_db_connection(profile, region):
    endpoints = describe_endpoints(profile, region)

    # Source endpoint
    db_engine = endpoints[0][2]
    host = endpoints[0][3]
    db = endpoints[0][4]
    port = endpoints[0][5]
    user = endpoints[0][6]

    # Fetch DB Password from AWS Secrets Manager
    password = ""

    return {
        "host": host,
        "port": port,
        "service": db,
        "user": "admin",
        "password": password,
    }


def get_target_db_connection(profile, region):
    endpoints = describe_endpoints(profile, region, print_result=False)

    # Target endpoint
    db_engine = endpoints[1][2]
    host = endpoints[1][3]
    db = endpoints[1][4]
    port = endpoints[1][5]
    user = endpoints[1][6]

    # Fetch DB Password from AWS Secrets Manager
    password = ""

    return {
        "host": host,
        "port": port,
        "service": db,
        "user": user,
        "password": password,
    }


def validate_source_target_structures(profile, region, table_name):
    """
    Validate Source and Target DB structures
    """
    schema = table_name.split(".")[0]
    table = table_name.split(".")[1]

    source_config = get_source_db_connection(profile, region)

    # Get Metadata from Source DB
    source_metadata = oracle_table_metadata(source_config, schema, table)

    target_config = get_target_db_connection(profile, region)

    # Get Metadata from Target DB
    target_metadata = postgres_table_metadata(target_config, schema, table)


    wb = openpyxl.Workbook()
    sheet = wb['Sheet']              # Default sheet name is 'Sheet'
    sheet.title = 'structure_comparison'
    sheet.sheet_properties.tabColor = "1072BA"

    current_time = (
            datetime.now()
            .strftime("%Y-%m-%d %H:%M")
            .replace(" ", "-")
            .replace(":", "-")
        )

    target_file = '../table_structure_comparison/structure_comparison' + '_' + current_time + '.xlsx'

    for i in range(len(source_metadata)):
        for j in range(len(source_metadata[i])):
            sheet.cell(row=i+1, column=j+1).value = source_metadata[i][j]
        
        k = len(source_metadata[i]) + 1

        for j in range(len(target_metadata[i])):
            sheet.cell(row=i+1, column= k + j + 1).value = target_metadata[i][j]        

        #sheet.cell(row=i, column=1).value = 'Hello World'

    wb.save(target_file)
