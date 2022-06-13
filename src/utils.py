import os
import textwrap
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles.borders import Border, Side
from tabulate import tabulate


def convert_schemas_to_lowercase():
    return {
        "rule-type": "transformation",
        "rule-id": "3",
        "rule-name": "convert-schemas-to-lower",
        "rule-action": "convert-lowercase",
        "rule-target": "schema",
        "object-locator": {"schema-name": "%"},
    }


def convert_tables_to_lowercase():
    return {
        "rule-type": "transformation",
        "rule-id": "4",
        "rule-name": "convert-tables-to-lower",
        "rule-action": "convert-lowercase",
        "rule-target": "table",
        "object-locator": {"schema-name": "%", "table-name": "%"},
    }


def convert_columns_to_lowercase():
    return {
        "rule-type": "transformation",
        "rule-id": "5",
        "rule-name": "convert-columns-to-lowercase",
        "rule-action": "convert-lowercase",
        "rule-target": "column",
        "object-locator": {"schema-name": "%", "table-name": "%", "column-name": "%"},
    }


def get_aws_cli_profile():
    # ----------------------------------------------------------------------------------------------#
    # Look for AWS CLI profiles in ~/.aws/config
    # ----------------------------------------------------------------------------------------------#
    aws_config_file_path = os.path.join(Path.home(), ".aws", "config")
    profiles = []

    if not os.path.exists(aws_config_file_path):
        return profiles

    with open(aws_config_file_path, "r") as aws_cli_profile_file:
        aws_cli_profile_list = aws_cli_profile_file.readlines()

        for line in aws_cli_profile_list:
            if "[" in line:
                aws_cli_profile_name = line.split("[")[1].split("]")[0]
                aws_cli_profile_name = aws_cli_profile_name.replace(
                    "profile", ""
                ).strip()
                profiles.append(aws_cli_profile_name)

    return profiles


def write_to_excel_file(list1, list2):
    """
    Writes the input M x N matrices to an Excel file.

    Assumption is that, both the lists are of the same size (i.e, each have
    M rows, and there are N cells in each row).

    :param list1: M x N matrix
    :param list2: M x N matrix

    :return: None
    """
    no_cells_in_each_row_in_list1 = set([len(row) for row in list1])
    no_cells_in_each_row_in_list2 = set([len(row) for row in list2])

    if (
        len(list1) != len(list2)
        or no_cells_in_each_row_in_list1 != no_cells_in_each_row_in_list2
    ):
        print("Input lists are not of the same size.")

        print(
            tabulate(
                list1[1:],
                headers=list1[0],
                tablefmt="fancy_grid",
            )
        )

        print(
            tabulate(
                list2[1:],
                headers=list2[0],
                tablefmt="fancy_grid",
            )
        )

        return

    wb = openpyxl.Workbook()
    sheet = wb["Sheet"]  # Default sheet name is 'Sheet'
    sheet.title = "structure_comparison"
    sheet.sheet_properties.tabColor = "1072BA"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    current_time = (
        datetime.now().strftime("%Y_%m_%d %H:%M").replace(" ", "_").replace(":", "_")
    )

    target_file = (
        "../table_structure_validation/structure_comparison"
        + "_"
        + current_time
        + ".xlsx"
    )

    for i in range(len(list1)):
        for j in range(len(list1[i])):
            sheet.cell(row=i + 1, column=j + 1).value = list1[i][j]
            sheet.cell(row=i + 1, column=j + 1).border = thin_border

        k = len(list1[i]) + 1

        for j in range(len(list2[i])):
            sheet.cell(row=i + 1, column=k + j + 1).value = list2[i][j]
            sheet.cell(row=i + 1, column=k + j + 1).border = thin_border

    wb.save(target_file)

    print(f"-> Data written to excel file: {target_file}")


def print_messages(messages, headers):
    """
    Prints the messages in the input list.

    :param messages: List of Lists
    :param headers: List of strings

    :return: None
    """
    wrapped_messages = []

    for message in messages:
        wrapped_message = "\n".join(
            textwrap.wrap(message[0], width=180, replace_whitespace=False)
        )

        wrapped_messages.append([wrapped_message])

    print(
        tabulate(
            wrapped_messages,
            headers=headers,
            tablefmt="fancy_grid",
        )
    )
